#!/usr/bin/env python3
"""
東海エアサービス 給与計算スクリプト（Lark API統合版）
使い方:
  python3 payroll_calc.py --month 2602                   # Lark APIから自動取得（デフォルト）
  python3 payroll_calc.py --month 2602 --api             # API強制（フォールバックなし）
  python3 payroll_calc.py --month 2602 --xlsx /path/to/flow.xlsx  # xlsx入力モード
  python3 payroll_calc.py --month 2602 --xlsx auto       # Downloadsから最新xlsx自動検出
  python3 payroll_calc.py --month 2602 --csv             # 従来のCSV入力モード
  python3 payroll_calc.py --month 2602 --dry-run         # API取得テスト（計算のみ）
  python3 payroll_calc.py --month 2602 --special '2026-02-10:単独撮影:15000'

データソース（優先順位）:
  1. Lark Attendance API (user_tasks/query) → 勤怠データ（デフォルト）
  2. --xlsx 指定時 or API失敗時 → Lark勤怠エクスポートxlsx
  3. Lark Approval API → 経費精算データ
  4. CRM受注台帳 → 撮影実績（単独撮影・加算判定）
  5. CSV（フォールバック / --csv 指定時）
"""

import csv
import glob as glob_mod
import json
import os
import sys
import argparse
import smtplib
import ssl
import time
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Google Sheets / PDF（オプショナル）
try:
    import gspread
    from google.oauth2.service_account import Credentials as SACredentials
    HAS_GSPREAD = True
except ImportError:
    HAS_GSPREAD = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as rl_canvas
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from dotenv import load_dotenv
    load_dotenv("/home/user/tokaiair/.env")
except ImportError:
    pass

# ==================== 設定 ====================
SCRIPT_DIR = Path(__file__).parent
BASE_DIR = Path("/home/user/tokaiair/payroll")

# Lark API設定
LARK_API_BASE = "https://open.larksuite.com/open-apis"
CRM_BASE_TOKEN = "BodWbgw6DaHP8FspBTYjT8qSpOe"
TABLE_ORDERS = "tbldLj2iMJYocct6"
TABLE_EXPENSE_LOG = "tbliYwPFbxxINAfk"  # 経費精算ログテーブル

# 対象社員
TARGET_EMPLOYEE = "新美 光"
TARGET_EMAIL = "h.niimi@tokaiair.com"
NIIMI_USER_ID = "11agc33c"  # Lark Attendance API用employee_id
NIIMI_OPEN_ID = "ou_189dc637b61a83b886d356becb3ae18e"

# 給与体系
FULL_DAY_HOURS_MIN = 6.0      # フル日の最低労働時間
HALF_DAY_HOURS_MIN = 4.0      # 半日の最低労働時間
FULL_DAY_WAGE = 16000          # フル日給
HALF_DAY_WAGE = 8000           # 半日給
VEHICLE_ALLOWANCE_FULL = 1000  # 車両手当（フル日）
VEHICLE_ALLOWANCE_HALF = 500   # 車両手当（半日）
GAS_RATE = 15                  # ガソリン代単価（円/km）
SOLO_SHOOTING_WAGE = 15000     # 単独撮影日給
POINT_CLOUD_WAGE = 10560       # 点群処理日給

# APIリトライ設定
API_MAX_RETRIES = 3
API_RETRY_DELAY = 5  # 秒

# Google Sheets設定
GOOGLE_SA_KEY_PATHS = [
    Path("/mnt/c/Users/USER/Documents/_data/drive-organizer-489313-9230cf87e259.json"),
    Path("/tmp/google_sa.json"),
]
SPREADSHEET_ID = "1dJ2Yx2heeRU9gUnrAv3jrO00zrjxmt6Fo2CaaKnwI_w"

# PDF生成設定
PDF_OUTPUT_DIR = Path(__file__).parent.parent / "payroll"
FONT_REGULAR_PATH = "/mnt/c/Windows/Fonts/BIZ-UDGothicR.ttc"
FONT_BOLD_PATH = "/mnt/c/Windows/Fonts/BIZ-UDGothicB.ttc"
FONT_REGULAR = "BIZUDGothic"
FONT_BOLD = "BIZUDGothicBold"

COMPANY_INFO = {
    "name": "東海エアサービス株式会社",
    "rep": "代表取締役 國本 洋輔",
    "zip": "〒465-0077",
    "addr": "愛知県名古屋市名東区植園町1-9-3 LM1205",
    "invoice_no": "T5180001140533",
    "email": "info@tokaiair.com",
    "url": "https://www.tokaiair.com",
}

# 給与計算テーブル（Lark Base）
TABLE_PAYROLL = "tbllGwzN1GWwdd4L"

# 國本のopen_id（確認DM送信先）
KUNIMOTO_OPEN_ID = "ou_d2e2e520a442224ea9d987c6186341ce"

# 扶養人数別源泉徴収税額表（甲欄・令和8年）
WITHHOLDING_TABLE = [
    (  88000,  89000, [  130,    0,    0,    0,    0,    0,    0]),
    (  89000,  90000, [  180,    0,    0,    0,    0,    0,    0]),
    (  90000,  91000, [  230,    0,    0,    0,    0,    0,    0]),
    (  91000,  92000, [  290,    0,    0,    0,    0,    0,    0]),
    (  92000,  93000, [  340,    0,    0,    0,    0,    0,    0]),
    (  93000,  94000, [  390,    0,    0,    0,    0,    0,    0]),
    (  94000,  95000, [  440,    0,    0,    0,    0,    0,    0]),
    (  95000,  96000, [  490,    0,    0,    0,    0,    0,    0]),
    (  96000,  97000, [  540,    0,    0,    0,    0,    0,    0]),
    (  97000,  98000, [  590,    0,    0,    0,    0,    0,    0]),
    (  98000,  99000, [  640,    0,    0,    0,    0,    0,    0]),
    (  99000, 101000, [  720,    0,    0,    0,    0,    0,    0]),
    ( 101000, 103000, [  830,    0,    0,    0,    0,    0,    0]),
    ( 103000, 105000, [  930,    0,    0,    0,    0,    0,    0]),
    ( 105000, 107000, [ 1030,    0,    0,    0,    0,    0,    0]),
    ( 107000, 109000, [ 1130,    0,    0,    0,    0,    0,    0]),
    ( 109000, 111000, [ 1240,    0,    0,    0,    0,    0,    0]),
    ( 111000, 113000, [ 1340,    0,    0,    0,    0,    0,    0]),
    ( 113000, 115000, [ 1440,    0,    0,    0,    0,    0,    0]),
    ( 115000, 117000, [ 1540,    0,    0,    0,    0,    0,    0]),
    ( 117000, 119000, [ 1640,    0,    0,    0,    0,    0,    0]),
    ( 119000, 121000, [ 1740,    0,    0,    0,    0,    0,    0]),
    ( 121000, 123000, [ 1840,    0,    0,    0,    0,    0,    0]),
    ( 123000, 125000, [ 1950,    0,    0,    0,    0,    0,    0]),
    ( 125000, 127000, [ 2050,    0,    0,    0,    0,    0,    0]),
    ( 127000, 129000, [ 2150,    0,    0,    0,    0,    0,    0]),
    ( 129000, 131000, [ 2250,    0,    0,    0,    0,    0,    0]),
    ( 131000, 133000, [ 2350,    0,    0,    0,    0,    0,    0]),
    ( 133000, 135000, [ 2450,    0,    0,    0,    0,    0,    0]),
    ( 135000, 137000, [ 2550,    0,    0,    0,    0,    0,    0]),
    ( 137000, 139000, [ 2650,    0,    0,    0,    0,    0,    0]),
    ( 139000, 141000, [ 2750,    0,    0,    0,    0,    0,    0]),
    ( 141000, 143000, [ 2850,    0,    0,    0,    0,    0,    0]),
    ( 143000, 145000, [ 2950,    0,    0,    0,    0,    0,    0]),
    ( 145000, 147000, [ 3050,    0,    0,    0,    0,    0,    0]),
    ( 147000, 149000, [ 3150,    0,    0,    0,    0,    0,    0]),
    ( 149000, 151000, [ 3250,    0,    0,    0,    0,    0,    0]),
    ( 151000, 153000, [ 3350,  1440,    0,    0,    0,    0,    0]),
    ( 153000, 155000, [ 3450,  1540,    0,    0,    0,    0,    0]),
    ( 155000, 157000, [ 3550,  1640,    0,    0,    0,    0,    0]),
    ( 157000, 160000, [ 3700,  1790,    0,    0,    0,    0,    0]),
    ( 160000, 163000, [ 3900,  1990,    0,    0,    0,    0,    0]),
    ( 163000, 166000, [ 4100,  1440,    0,    0,    0,    0,    0]),
    ( 166000, 169000, [ 4300,  1640,    0,    0,    0,    0,    0]),
    ( 169000, 172000, [ 4500,  1840,    0,    0,    0,    0,    0]),
    ( 172000, 175000, [ 4700,  2040,    0,    0,    0,    0,    0]),
    ( 175000, 178000, [ 4900,  2240,    0,    0,    0,    0,    0]),
    ( 178000, 181000, [ 5100,  2440,    0,    0,    0,    0,    0]),
    ( 181000, 184000, [ 5300,  2640,    0,    0,    0,    0,    0]),
    ( 184000, 187000, [ 5500,  2840,  230,    0,    0,    0,    0]),
    ( 187000, 190000, [ 5700,  3040,  430,    0,    0,    0,    0]),
    ( 190000, 193000, [ 5900,  3240,  630,    0,    0,    0,    0]),
    ( 193000, 196000, [ 6100,  3440,  830,    0,    0,    0,    0]),
    ( 196000, 199000, [ 6300,  3640, 1030,    0,    0,    0,    0]),
    ( 199000, 202000, [ 6500,  3840, 1230,    0,    0,    0,    0]),
    ( 202000, 205000, [ 6700,  4040, 1430,    0,    0,    0,    0]),
    ( 205000, 210000, [ 6950,  4290, 1680,    0,    0,    0,    0]),
    ( 210000, 215000, [ 7350,  4690, 2080,    0,    0,    0,    0]),
    ( 215000, 220000, [ 7750,  5090, 2480,    0,    0,    0,    0]),
    ( 220000, 225000, [ 8150,  5490, 2880,  280,    0,    0,    0]),
    ( 225000, 230000, [ 8550,  5890, 3280,  680,    0,    0,    0]),
    ( 230000, 235000, [ 8950,  6290, 3680, 1080,    0,    0,    0]),
    ( 235000, 240000, [ 9350,  6690, 4080, 1480,    0,    0,    0]),
    ( 240000, 245000, [ 9750,  7090, 4480, 1880,    0,    0,    0]),
    ( 245000, 250000, [10150,  7490, 4880, 2280,    0,    0,    0]),
    ( 250000, 255000, [10550,  7890, 5280, 2680,   70,    0,    0]),
]


# ==================== Config読み込み ====================
def load_config():
    """automation_config.jsonから認証情報を読み込み"""
    for p in [
        Path("/mnt/c/Users/USER/Documents/_data/automation_config.json"),
        SCRIPT_DIR / "automation_config.json",
    ]:
        if p.exists():
            with open(p) as f:
                cfg = json.load(f)
            if not str(cfg.get("lark", {}).get("app_id", "")).startswith("${"):
                return cfg
    # env vars fallback
    return {
        "lark": {
            "app_id": os.environ.get("LARK_APP_ID", ""),
            "app_secret": os.environ.get("LARK_APP_SECRET", ""),
        }
    }


# ==================== Lark API共通 ====================
def lark_api_request(url, method="GET", data=None, token=None, retries=API_MAX_RETRIES):
    """Lark APIリクエスト（リトライ付き）"""
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    encoded_data = json.dumps(data).encode() if data else None

    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, data=encoded_data, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=30) as r:
                result = json.loads(r.read())
                if result.get("code", 0) != 0:
                    msg = result.get("msg", "unknown error")
                    print(f"  API error (code={result.get('code')}): {msg}")
                    if attempt < retries - 1:
                        time.sleep(API_RETRY_DELAY)
                        continue
                return result
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            print(f"  API request failed (attempt {attempt+1}/{retries}): {e}")
            if attempt < retries - 1:
                time.sleep(API_RETRY_DELAY)
            else:
                raise
    return None


def lark_get_token(config):
    """tenant_access_tokenを取得"""
    app_id = config["lark"]["app_id"]
    app_secret = config["lark"]["app_secret"]
    result = lark_api_request(
        f"{LARK_API_BASE}/auth/v3/tenant_access_token/internal",
        method="POST",
        data={"app_id": app_id, "app_secret": app_secret},
    )
    return result["tenant_access_token"]


def lark_get_user_id(token, email):
    """メールアドレスからuser_idを取得"""
    result = lark_api_request(
        f"{LARK_API_BASE}/contact/v3/users/batch_get_id?user_id_type=user_id",
        method="POST",
        data={"emails": [email]},
        token=token,
    )
    users = result.get("data", {}).get("user_list", [])
    if users and users[0].get("user_id"):
        return users[0]["user_id"]
    return None


# ==================== Lark勤怠API ====================
def fetch_attendance_from_api(token, user_id, year: int, month: int) -> list[dict]:
    """
    Lark Attendance API user_tasks/query で月次の打刻記録を取得。
    各日の check_in_record / check_out_record からUnixタイムスタンプを読み取り、
    parse_attendance_csv() と同じ出力フォーマットに変換する。
    打刻なし（N/A）の日はスキップ。
    """
    # 対象月の開始日・終了日（YYYYMMDD整数）
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)

    check_date_from = int(start_date.strftime("%Y%m%d"))
    check_date_to = int(end_date.strftime("%Y%m%d"))

    print(f"  user_tasks/query: user_id={user_id}, {check_date_from}-{check_date_to}")

    result = lark_api_request(
        f"{LARK_API_BASE}/attendance/v1/user_tasks/query?employee_type=employee_id",
        method="POST",
        data={
            "user_ids": [user_id],
            "check_date_from": check_date_from,
            "check_date_to": check_date_to,
        },
        token=token,
    )

    if not result or result.get("code") != 0:
        code = result.get("code", "N/A") if result else "N/A"
        msg = result.get("msg", "") if result else "no response"
        print(f"  user_tasks/query failed (code={code}): {msg}")
        return []

    records = []
    user_task_results = result.get("data", {}).get("user_task_results", [])

    for task in user_task_results:
        # 日付
        date_val = task.get("day", 0)
        if not date_val:
            continue
        date_str_raw = str(date_val)
        if len(date_str_raw) == 8:
            date_str = f"{date_str_raw[:4]}-{date_str_raw[4:6]}-{date_str_raw[6:8]}"
        else:
            continue

        # 出勤打刻
        check_in = task.get("check_in_record", {})
        check_out = task.get("check_out_record", {})

        # check_time が "0" や空の場合は打刻なし
        in_time_str = check_in.get("check_time", "0") if check_in else "0"
        out_time_str = check_out.get("check_time", "0") if check_out else "0"

        # Unixタイムスタンプ（秒）を変換
        try:
            in_ts = int(in_time_str)
            out_ts = int(out_time_str)
        except (ValueError, TypeError):
            continue

        # 打刻なしの日はスキップ
        if in_ts == 0 or out_ts == 0:
            if in_ts != 0 and out_ts == 0:
                # 出勤のみ（退勤打刻漏れ）
                t_in = datetime.fromtimestamp(in_ts)
                records.append({
                    "date": date_str,
                    "time_in": t_in.strftime("%Y-%m-%d %H:%M:%S"),
                    "time_out": "",
                    "worked_hours": 0,
                    "note": "退勤打刻なし",
                })
            continue

        t_in = datetime.fromtimestamp(in_ts)
        t_out = datetime.fromtimestamp(out_ts)
        worked_hours = (t_out - t_in).total_seconds() / 3600 - 1.0  # 休憩1h控除

        records.append({
            "date": date_str,
            "time_in": t_in.strftime("%Y-%m-%d %H:%M:%S"),
            "time_out": t_out.strftime("%Y-%m-%d %H:%M:%S"),
            "worked_hours": round(max(worked_hours, 0), 2),
        })

    return sorted(records, key=lambda x: x["date"])


# ==================== Lark経費精算（Bitable経費精算ログ） ====================
# Approval APIのバグ回避: 経費精算ログテーブル（Bitable）から取得
# Lark Base Automationで承認時に自動レコード追加 → このテーブルから全件取得可能
EXPENSE_APPROVAL_CODE = "25922894-E416-4D50-90E6-EFAF8D88DDC1"  # 参照用に残す

def fetch_expenses_from_api(token, user_id, year: int, month: int) -> dict:
    """
    経費精算ログテーブル（Bitable API）から対象月の経費データを取得。
    Approval APIのバグ（instances/listが0件、queryが最新10件のみ）を回避。
    対象月フィールド（YYMM形式）でフィルタして該当月のレコードのみ取得。
    """
    target_month = f"{year % 100:02d}{month:02d}"  # YYMM形式
    print(f"  経費精算ログテーブル: {TABLE_EXPENSE_LOG}")
    print(f"  対象月: {target_month}")

    gas_km = 0
    fixed_expenses = []
    special_dates = []

    # Bitable records/search APIで対象月フィルタ
    page_token = None
    while True:
        search_body = {
            "filter": {
                "conjunction": "and",
                "conditions": [
                    {
                        "field_name": "対象月",
                        "operator": "is",
                        "value": [target_month],
                    },
                    {
                        "field_name": "ステータス",
                        "operator": "is",
                        "value": ["承認済み"],
                    },
                ],
            },
            "page_size": 500,
        }
        if page_token:
            search_body["page_token"] = page_token

        result = lark_api_request(
            f"{LARK_API_BASE}/bitable/v1/apps/{CRM_BASE_TOKEN}/tables/{TABLE_EXPENSE_LOG}/records/search?user_id_type=open_id",
            method="POST",
            data=search_body,
            token=token,
        )

        if not result or result.get("code") != 0:
            # records/search失敗時はGET全件取得にフォールバック
            print(f"  records/search失敗、GET全件取得にフォールバック")
            return _fetch_expenses_bitable_fallback(token, target_month)

        data = result.get("data", {})
        items = data.get("items", [])
        print(f"  取得レコード数: {len(items)}")

        for rec in items:
            fields = rec.get("fields", {})
            parsed = _parse_expense_log_record(fields)
            gas_km += parsed.get("gas_km", 0)
            fixed_expenses.extend(parsed.get("fixed_expenses", []))
            special_dates.extend(parsed.get("special_dates", []))

        if not data.get("has_more"):
            break
        page_token = data.get("page_token")
        time.sleep(0.3)

    print(f"  走行距離合計: {gas_km}km / ガソリン代: {gas_km * GAS_RATE:,}円")
    if fixed_expenses:
        print(f"  実費精算: {len(fixed_expenses)}件")
        for e in fixed_expenses:
            print(f"    {e.get('date','')} {e.get('content','')}: {e.get('amount',0):,}円")

    return {
        "gas_km": gas_km,
        "fixed_expenses": fixed_expenses,
        "special_dates": list(set(special_dates)),
    }


def _fetch_expenses_bitable_fallback(token, target_month):
    """records/search失敗時のGET全件取得フォールバック"""
    gas_km = 0
    fixed_expenses = []
    special_dates = []

    page_token = None
    while True:
        url = f"{LARK_API_BASE}/bitable/v1/apps/{CRM_BASE_TOKEN}/tables/{TABLE_EXPENSE_LOG}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"

        result = lark_api_request(url, token=token)
        if not result or result.get("code") != 0:
            break

        data = result.get("data", {})
        items = data.get("items", [])

        for rec in items:
            fields = rec.get("fields", {})
            # 対象月フィルタ（クライアント側）
            rec_month = _extract_text(fields.get("対象月", ""))
            if rec_month != target_month:
                continue
            # ステータスフィルタ
            status = _extract_text(fields.get("ステータス", ""))
            if status and status != "承認済み":
                continue

            parsed = _parse_expense_log_record(fields)
            gas_km += parsed.get("gas_km", 0)
            fixed_expenses.extend(parsed.get("fixed_expenses", []))
            special_dates.extend(parsed.get("special_dates", []))

        if not data.get("has_more"):
            break
        page_token = data.get("page_token")
        time.sleep(0.3)

    return {
        "gas_km": gas_km,
        "fixed_expenses": fixed_expenses,
        "special_dates": list(set(special_dates)),
    }


def _parse_expense_log_record(fields) -> dict:
    """経費精算ログテーブルの1レコードを解析"""
    gas_km = 0
    fixed_expenses = []
    special_dates = []

    # 距離km（数値フィールド）
    distance = fields.get("距離km")
    if distance is not None:
        try:
            km = int(float(str(distance)))
            if km > 0:
                gas_km = km
        except (ValueError, TypeError):
            pass

    # 金額（実費：数値フィールド）
    amount = fields.get("金額")
    if amount is not None:
        try:
            amt = int(float(str(amount)))
            if amt > 0:
                # 日付フィールドから日付文字列を取得
                date_val = fields.get("日付")
                date_str = ""
                if isinstance(date_val, (int, float)):
                    date_str = datetime.fromtimestamp(date_val / 1000).strftime("%Y-%m-%d")
                elif isinstance(date_val, str):
                    date_str = date_val[:10]

                content = _extract_text(fields.get("内容", ""))
                exp_type = _extract_text(fields.get("経費タイプ", ""))

                fixed_expenses.append({
                    "date": date_str,
                    "type": exp_type or "実費",
                    "content": content or "経費精算",
                    "amount": amt,
                })
        except (ValueError, TypeError):
            pass

    # 撮影キーワード検出
    content = _extract_text(fields.get("内容", "")) + _extract_text(fields.get("理由", ""))
    keywords = ["撮影", "空撮", "測量", "計測", "ドローン"]
    if any(kw in content for kw in keywords):
        date_val = fields.get("日付")
        if isinstance(date_val, (int, float)):
            d = datetime.fromtimestamp(date_val / 1000).strftime("%Y-%m-%d")
            special_dates.append(d)

    return {
        "gas_km": gas_km,
        "fixed_expenses": fixed_expenses,
        "special_dates": special_dates,
    }


# ==================== CRM受注台帳から撮影実績 ====================
def fetch_shooting_records(token, year: int, month: int) -> list[dict]:
    """
    受注台帳から対象月の新美担当案件を取得
    単独撮影・複数現場加算の自動判定
    """
    all_records = []
    page_token = None

    while True:
        url = f"{LARK_API_BASE}/bitable/v1/apps/{CRM_BASE_TOKEN}/tables/{TABLE_ORDERS}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"

        result = lark_api_request(url, token=token)
        if not result or result.get("code") != 0:
            break

        data = result.get("data", {})
        items = data.get("items", [])
        all_records.extend(items)

        if not data.get("has_more"):
            break
        page_token = data.get("page_token")
        time.sleep(0.3)

    # 対象月の範囲
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1) - timedelta(days=1)

    shooting_records = []
    date_counts = {}  # 日別の案件数（複数現場判定用）

    for rec in all_records:
        fields = rec.get("fields", {})

        # 担当営業チェック
        rep = _extract_person_name(fields.get("担当営業", "")) or _extract_person_name(fields.get("担当", ""))
        if "新美" not in str(rep):
            continue

        # 撮影日 or 納品日を探す
        shoot_date = None
        for key in ("撮影日", "実施日", "作業日", "納品日"):
            val = fields.get(key)
            if val:
                try:
                    if isinstance(val, (int, float)):
                        shoot_date = datetime.fromtimestamp(val / 1000)
                    elif isinstance(val, str):
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%dT%H:%M:%S"):
                            try:
                                shoot_date = datetime.strptime(val.split("+")[0].split("T")[0], fmt.split("T")[0])
                                break
                            except ValueError:
                                continue
                except (ValueError, OSError, TypeError):
                    continue
                if shoot_date:
                    break

        if not shoot_date:
            continue

        if not (start_date <= shoot_date <= end_date):
            continue

        date_key = shoot_date.strftime("%Y-%m-%d")
        case_name = _extract_text(fields.get("案件名", ""))
        product = _extract_text(fields.get("商材", "")) or _extract_text(fields.get("案件種別", ""))

        # 日別カウント
        date_counts[date_key] = date_counts.get(date_key, 0) + 1

        shooting_records.append({
            "date": date_key,
            "case_name": case_name,
            "product": product,
            "fields": fields,
        })

    # 単独撮影判定: その日に新美が1件だけ担当 = 単独撮影の可能性
    for rec in shooting_records:
        date = rec["date"]
        rec["is_solo"] = date_counts.get(date, 0) == 1
        rec["same_day_count"] = date_counts.get(date, 0)

    return shooting_records


def _extract_text(val):
    """Larkフィールド値からテキストを抽出"""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, list):
        texts = []
        for item in val:
            if isinstance(item, dict):
                texts.append(item.get("text", "") or item.get("name", "") or str(item.get("value", "")))
            elif isinstance(item, str):
                texts.append(item)
        return ", ".join(t for t in texts if t)
    if isinstance(val, dict):
        return val.get("text", "") or val.get("name", "") or val.get("value", "")
    return str(val) if val else ""


def _extract_person_name(val):
    """Lark人物フィールドから名前を抽出"""
    if isinstance(val, list):
        names = []
        for item in val:
            if isinstance(item, dict):
                names.append(item.get("name", "") or item.get("en_name", ""))
        return ", ".join(n for n in names if n)
    if isinstance(val, dict):
        return val.get("name", "") or val.get("en_name", "")
    if isinstance(val, str):
        return val
    return ""


# ==================== CSV読み込み（フォールバック用、既存ロジック維持） ====================
def parse_attendance_csv(filepath: Path) -> list[dict]:
    """
    Lark勤怠CSVを読み込んで日別勤務記録を返す
    各日のout/inペアを合わせて実働時間を計算
    """
    records = []
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    by_date = {}
    for row in rows:
        name = row.get("氏名", "").strip()
        if not name or name != TARGET_EMPLOYEE:
            continue
        date_str = row.get("日付", "").strip()
        punch_time = row.get("打刻時間", "").strip()
        result = row.get("打刻結果", "").strip()

        if "更新されています" in result:
            continue

        if date_str not in by_date:
            by_date[date_str] = []
        by_date[date_str].append(punch_time)

    for date_str, times in by_date.items():
        if len(times) < 2:
            continue
        times_sorted = sorted(times)
        t_in = datetime.strptime(times_sorted[0], "%Y-%m-%d %H:%M:%S")
        t_out = datetime.strptime(times_sorted[-1], "%Y-%m-%d %H:%M:%S")
        worked_hours = (t_out - t_in).total_seconds() / 3600 - 1.0

        records.append({
            "date": date_str,
            "time_in": times_sorted[0],
            "time_out": times_sorted[-1],
            "worked_hours": round(worked_hours, 2),
        })

    return sorted(records, key=lambda x: x["date"])


def parse_expenses_csv(filepath: Path) -> dict:
    """
    Lark経費精算CSVを読み込んで集計
    """
    gas_km = 0
    fixed_expenses = []
    special_dates = []

    if not filepath.exists():
        return {"gas_km": 0, "fixed_expenses": [], "special_dates": []}

    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    for row in rows:
        applicant = row.get("申請者の名前", "").strip()
        if applicant != TARGET_EMPLOYEE:
            continue
        status = row.get("ステータス", "").strip()
        if status not in ["承認済み"]:
            continue

        exp_type = row.get("経費タイプ", "").strip()
        content = row.get("内容", "").strip()
        date_str = row.get("日付（年ー月ー日）", "").strip()
        amount_str = row.get("金額", "").strip().replace(",", "")
        distance_str = row.get("距離合計", "").strip()

        if distance_str and distance_str.isdigit():
            km = int(distance_str)
            if km > 0:
                gas_km += km

        if amount_str:
            try:
                amount = int(float(amount_str))
                if amount > 0:
                    fixed_expenses.append({
                        "date": date_str,
                        "type": exp_type,
                        "content": content or exp_type,
                        "amount": amount,
                    })
            except ValueError:
                pass

        keywords = ["撮影", "空撮", "測量", "計測", "ドローン"]
        reason = row.get("経費精算の理由", "") + content
        if any(kw in reason for kw in keywords):
            if date_str and date_str not in special_dates:
                special_dates.append(date_str)

    return {
        "gas_km": gas_km,
        "fixed_expenses": fixed_expenses,
        "special_dates": special_dates,
    }


# ==================== xlsx読み込み（Lark勤怠エクスポート） ====================
DOWNLOADS_DIRS = [
    Path("/mnt/c/Users/USER/Downloads"),
    Path.home() / "Downloads",
]


def find_latest_attendance_xlsx(month: str) -> Path | None:
    """
    Downloadsフォルダからflow_YYYYMMDD_YYYYMMDD.xlsxを自動検出。
    --month 2602 → 2026年2月 → flow_20260201_20260228.xlsx を探す。
    見つからなければ最新のflow_*.xlsxを返す。
    """
    year = 2000 + int(month[:2])
    mon = int(month[2:])
    # 対象月のファイル名パターン
    prefix = f"flow_{year}{mon:02d}"

    for dl_dir in DOWNLOADS_DIRS:
        if not dl_dir.exists():
            continue

        # 完全一致を優先
        candidates = sorted(dl_dir.glob(f"{prefix}*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if candidates:
            return candidates[0]

        # フォールバック: 最新のflow_*.xlsx
        all_flows = sorted(dl_dir.glob("flow_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if all_flows:
            return all_flows[0]

    return None


def parse_attendance_xlsx(filepath: Path, employee_name: str = None) -> list[dict]:
    """
    Lark勤怠管理コンソールからエクスポートしたxlsxを読み込み、
    parse_attendance_csv()と同じ出力フォーマットに変換する。

    xlsxフォーマット:
      氏名/部門/社員番号/勤怠管理グループ/社員タイプ/日付/曜日/シフト/打刻時間/打刻結果/...
      - 出勤行: シフト列が "YYYY-MM-DD 00:00:00"
      - 退勤行: シフト列が "YYYY-MM-DD 23:59:00"
      - 同じ日付で2行1組
    """
    try:
        import openpyxl
    except ImportError:
        print("  openpyxlが必要です: pip install openpyxl")
        sys.exit(1)

    employee_name = employee_name or TARGET_EMPLOYEE
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ヘッダー行を読み取り
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c) if c else "" for c in row]
        break

    # カラムインデックス特定
    col_map = {}
    for i, h in enumerate(headers):
        if h == "氏名":
            col_map["name"] = i
        elif h == "日付":
            col_map["date"] = i
        elif h == "シフト":
            col_map["shift"] = i
        elif h == "打刻時間":
            col_map["punch_time"] = i
        elif h == "打刻結果":
            col_map["punch_result"] = i

    required = ["name", "date", "shift", "punch_time"]
    missing = [k for k in required if k not in col_map]
    if missing:
        print(f"  xlsx必須カラムが見つかりません: {missing}")
        print(f"  検出されたヘッダー: {headers}")
        sys.exit(1)

    # データ行を読み取り、日付ごとにグループ化
    by_date = {}  # date_str -> {"in": datetime, "out": datetime}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row)
        name = str(cells[col_map["name"]] or "").strip()
        if name != employee_name:
            continue

        date_val = cells[col_map["date"]]
        shift_val = cells[col_map["shift"]]
        punch_val = cells[col_map["punch_time"]]

        if not date_val or not punch_val:
            continue

        # 日付を文字列に正規化
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).strip()[:10]

        # シフトで出勤/退勤を判定
        if isinstance(shift_val, datetime):
            shift_time = shift_val.strftime("%H:%M")
        else:
            shift_str = str(shift_val).strip()
            # "YYYY-MM-DD HH:MM:SS" 形式からHH:MMを抽出
            if " " in shift_str:
                shift_time = shift_str.split(" ")[1][:5]
            else:
                shift_time = shift_str[:5]

        # 打刻時間をdatetimeに変換
        if isinstance(punch_val, datetime):
            punch_dt = punch_val
        else:
            punch_str = str(punch_val).strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    punch_dt = datetime.strptime(punch_str, fmt)
                    break
                except ValueError:
                    continue
            else:
                continue

        if date_str not in by_date:
            by_date[date_str] = {}

        if shift_time == "00:00":
            by_date[date_str]["in"] = punch_dt
        elif shift_time == "23:59":
            by_date[date_str]["out"] = punch_dt
        else:
            # 不明なシフト → 時刻の早い方をin、遅い方をoutとして追加
            if "in" not in by_date[date_str]:
                by_date[date_str]["in"] = punch_dt
            else:
                by_date[date_str]["out"] = punch_dt

    wb.close()

    # parse_attendance_csv()と同じ出力フォーマットに変換
    records = []
    for date_str in sorted(by_date.keys()):
        entry = by_date[date_str]
        t_in = entry.get("in")
        t_out = entry.get("out")

        if t_in and t_out:
            worked_hours = (t_out - t_in).total_seconds() / 3600 - 1.0  # 休憩1h控除
            records.append({
                "date": date_str,
                "time_in": t_in.strftime("%Y-%m-%d %H:%M:%S"),
                "time_out": t_out.strftime("%Y-%m-%d %H:%M:%S"),
                "worked_hours": round(worked_hours, 2),
            })
        elif t_in:
            records.append({
                "date": date_str,
                "time_in": t_in.strftime("%Y-%m-%d %H:%M:%S"),
                "time_out": "",
                "worked_hours": 0,
                "note": "退勤打刻なし",
            })

    return records


# ==================== 源泉徴収 ====================
def get_withholding_tax(taxable_amount: int, dependents: int) -> int:
    """源泉徴収税額を計算（甲欄）"""
    dep_idx = min(dependents, 6)
    for (low, high, taxes) in WITHHOLDING_TABLE:
        if low <= taxable_amount < high:
            return taxes[dep_idx]
    if taxable_amount < 88000:
        return 0
    return 0


def classify_day(worked_hours: float) -> str:
    """労働時間から勤務区分を判定"""
    if worked_hours >= FULL_DAY_HOURS_MIN:
        return "フル日"
    elif worked_hours >= HALF_DAY_HOURS_MIN:
        return "半日"
    else:
        return "要確認"


# ==================== 給与計算メイン ====================
def calculate_payroll(month: str, dependents: int = 1, special_overrides: dict = None,
                      use_csv: bool = False, xlsx_path: str = None,
                      dry_run: bool = False, force_api: bool = False) -> dict:
    """
    給与計算メイン関数
    month: "2602" 形式
    use_csv: Trueの場合はCSV入力モード
    xlsx_path: xlsxファイルパス or "auto"（Downloadsから自動検出）
    dry_run: Trueの場合はAPI取得テスト（計算のみ）
    force_api: Trueの場合はAPI強制（フォールバックしない）
    """
    year = 2000 + int(month[:2])
    mon = int(month[2:])
    special_overrides = special_overrides or {}

    attendance = None
    expenses = None
    shooting_records = []

    # === xlsx入力モード ===
    if xlsx_path:
        if xlsx_path == "auto":
            resolved = find_latest_attendance_xlsx(month)
            if not resolved:
                print(f"  Downloadsフォルダにflow_*.xlsxが見つかりません")
                sys.exit(1)
            print(f"\n[XLSX] 自動検出: {resolved}")
        else:
            resolved = Path(xlsx_path)
            if not resolved.exists():
                print(f"  xlsxファイルが見つかりません: {resolved}")
                sys.exit(1)
            print(f"\n[XLSX] 読み込み: {resolved}")

        attendance = parse_attendance_xlsx(resolved)
        print(f"  勤怠レコード: {len(attendance)}件")
        for rec in attendance:
            hours = rec.get("worked_hours", 0)
            day_type = classify_day(hours)
            t_in = rec.get("time_in", "")[11:16] if rec.get("time_in") else "N/A"
            t_out = rec.get("time_out", "")[11:16] if rec.get("time_out") else "N/A"
            print(f"    {rec['date']} | {t_in} - {t_out} | {hours:.1f}h → {day_type}")

        # xlsxモードではexpensesはCSVフォールバックまたは空
        work_dir = BASE_DIR / month
        expenses_file = work_dir / "expenses.csv"
        if expenses_file.exists():
            print(f"[CSV] 経費データ読み込み: {expenses_file}")
            expenses = parse_expenses_csv(expenses_file)
        else:
            expenses = {"gas_km": 0, "fixed_expenses": [], "special_dates": []}
            print(f"  経費データなし（{expenses_file}）→ 経費精算は手動確認")

    if not use_csv and attendance is None:
        # === Lark API取得モード ===
        try:
            print(f"\n[API] Lark APIからデータ取得中...")
            config = load_config()
            token = lark_get_token(config)
            print(f"  tenant_access_token取得OK")

            # user_id: 定数を使用（API lookup不要）
            user_id = NIIMI_USER_ID
            print(f"  user_id: {user_id} (定数)")

            # 1. 勤怠データ取得（user_tasks/query）
            print(f"\n[API] 勤怠データ取得中 ({year}/{mon:02d})...")
            attendance = fetch_attendance_from_api(token, user_id, year, mon)
            if attendance:
                print(f"  勤怠レコード: {len(attendance)}件")
                if dry_run:
                    print(f"\n  [dry-run] 勤怠データ:")
                    for rec in attendance:
                        t_in = rec.get('time_in', 'N/A')[11:16] if rec.get('time_in') else 'N/A'
                        t_out = rec.get('time_out', 'N/A')[11:16] if rec.get('time_out') else 'N/A'
                        note = f" ({rec['note']})" if rec.get('note') else ""
                        print(f"    {rec['date']} | {t_in} - {t_out} | {rec.get('worked_hours',0):.1f}h{note}")
            else:
                print(f"  勤怠API取得: 0件")
                if force_api:
                    print(f"  [--api] API強制モード: フォールバックなし")
                else:
                    print(f"  xlsxフォールバックを試みます...")
                    # xlsxフォールバック
                    resolved = find_latest_attendance_xlsx(month)
                    if resolved:
                        print(f"\n[XLSX] 自動検出: {resolved}")
                        attendance = parse_attendance_xlsx(resolved)
                        if attendance:
                            print(f"  勤怠レコード: {len(attendance)}件")
                    if not attendance:
                        print(f"  xlsxフォールバック: ファイルなし → CSVフォールバック")

            # 2. 経費精算データ取得
            print(f"\n[API] 経費精算データ取得中 ({year}/{mon:02d})...")
            expenses = fetch_expenses_from_api(token, user_id, year, mon)
            if expenses and (expenses["gas_km"] > 0 or expenses["fixed_expenses"]):
                print(f"  走行距離: {expenses['gas_km']}km")
                print(f"  実費精算: {len(expenses['fixed_expenses'])}件")
                if dry_run:
                    print(f"\n  [dry-run] 経費データ:")
                    print(f"    ガソリン: {expenses['gas_km']}km x {GAS_RATE}円 = {expenses['gas_km'] * GAS_RATE:,}円")
                    for e in expenses["fixed_expenses"]:
                        print(f"    {e['date']} {e['content']}: {e['amount']:,}円")
            else:
                print(f"  経費API取得: データなし（CSVフォールバック検討）")

            # 3. CRM受注台帳から撮影実績
            print(f"\n[API] CRM受注台帳から撮影実績取得中...")
            shooting_records = fetch_shooting_records(token, year, mon)
            if shooting_records:
                print(f"  撮影実績: {len(shooting_records)}件")
                if dry_run:
                    print(f"\n  [dry-run] 撮影実績:")
                    for rec in shooting_records:
                        solo = "単独" if rec["is_solo"] else f"同日{rec['same_day_count']}件"
                        print(f"    {rec['date']} | {rec['case_name'][:25]} | {solo}")
            else:
                print(f"  撮影実績: 0件")

        except Exception as e:
            print(f"\n[API] API取得エラー: {e}")
            if force_api:
                print(f"  [--api] API強制モード: フォールバックなし")
                raise
            print(f"[API] xlsxフォールバックを試みます...")
            attendance = None
            expenses = None
            # xlsxフォールバック
            resolved = find_latest_attendance_xlsx(month)
            if resolved:
                print(f"\n[XLSX] 自動検出: {resolved}")
                attendance = parse_attendance_xlsx(resolved)
                if attendance:
                    print(f"  勤怠レコード: {len(attendance)}件")

    # === CSVフォールバック ===
    if attendance is None:
        work_dir = BASE_DIR / month
        attendance_file = work_dir / "attendance.csv"

        if not attendance_file.exists():
            if use_csv:
                print(f"  勤怠ファイルが見つかりません: {attendance_file}")
                sys.exit(1)
            else:
                print(f"  CSVフォールバック: 勤怠ファイルなし ({attendance_file})")
                print(f"  勤怠データがありません。APIまたはCSVを確認してください。")
                sys.exit(1)

        print(f"\n[CSV] 勤怠データ読み込み: {attendance_file}")
        attendance = parse_attendance_csv(attendance_file)
        print(f"  レコード数: {len(attendance)}")

    if expenses is None:
        work_dir = BASE_DIR / month
        expenses_file = work_dir / "expenses.csv"
        print(f"[CSV] 経費データ読み込み: {expenses_file}")
        expenses = parse_expenses_csv(expenses_file)

    # ==================== 勤務日分類 ====================
    print(f"\n{'='*60}")
    print(f"  東海エアサービス 給与計算 {month[:2]}年{month[2:]}月度")
    print(f"{'='*60}")

    # 撮影実績から特約業務を自動設定（手動overrideが優先）
    shooting_dates = {r["date"]: r for r in shooting_records}

    work_days = []
    for rec in attendance:
        date = rec["date"]
        hours = rec.get("worked_hours", 0)
        day_type = classify_day(hours)

        # 特約業務の上書き（手動指定が最優先）
        if date in special_overrides:
            override = special_overrides[date]
            day_type = override["type"]
            wage = override["amount"]
            vehicle = VEHICLE_ALLOWANCE_HALF if "半日" in day_type or hours < FULL_DAY_HOURS_MIN else VEHICLE_ALLOWANCE_FULL
        elif date in shooting_dates and shooting_dates[date]["is_solo"]:
            # CRM受注台帳から単独撮影と判定された場合
            day_type = "単独撮影"
            wage = SOLO_SHOOTING_WAGE
            vehicle = VEHICLE_ALLOWANCE_FULL
        elif day_type == "フル日":
            wage = FULL_DAY_WAGE
            vehicle = VEHICLE_ALLOWANCE_FULL
        elif day_type == "半日":
            wage = HALF_DAY_WAGE
            vehicle = VEHICLE_ALLOWANCE_HALF
        else:
            wage = 0
            vehicle = 0

        work_days.append({
            "date": date,
            "time_in": rec.get("time_in", "")[11:16] if rec.get("time_in") else "",
            "time_out": rec.get("time_out", "")[11:16] if rec.get("time_out") else "",
            "worked_hours": hours,
            "day_type": day_type,
            "wage": wage,
            "vehicle": vehicle,
        })

    # 打刻なし特約業務（出張申請ベース等）を追加
    attendance_dates = {rec["date"] for rec in attendance}
    for date, override in sorted(special_overrides.items()):
        if date not in attendance_dates:
            day_type = override["type"]
            wage = override["amount"]
            vehicle = VEHICLE_ALLOWANCE_FULL if "半日" not in day_type else VEHICLE_ALLOWANCE_HALF
            work_days.append({
                "date": date,
                "time_in": "",
                "time_out": "",
                "worked_hours": 0,
                "day_type": day_type,
                "wage": wage,
                "vehicle": vehicle,
                "note": "打刻なし（出張申請ベース）",
            })
    work_days.sort(key=lambda d: d["date"])

    # ==================== 集計 ====================
    base_wage = sum(d["wage"] for d in work_days)
    vehicle_total = sum(d["vehicle"] for d in work_days)
    gas_amount = expenses["gas_km"] * GAS_RATE
    fixed_total = sum(e["amount"] for e in expenses["fixed_expenses"])

    taxable = base_wage + vehicle_total
    withholding = get_withholding_tax(taxable, dependents)
    gross = base_wage + vehicle_total + gas_amount + fixed_total
    net = gross - withholding

    year_str = "20" + month[:2]
    month_str = month[2:]
    pay_date = f"{year_str}年{month_str}月15日（翌月15日払い）"

    result = {
        "month": month,
        "pay_date": pay_date,
        "dependents": dependents,
        "work_days": work_days,
        "base_wage": base_wage,
        "vehicle_total": vehicle_total,
        "gas_km": expenses["gas_km"],
        "gas_amount": gas_amount,
        "fixed_expenses": expenses["fixed_expenses"],
        "fixed_total": fixed_total,
        "taxable": taxable,
        "withholding": withholding,
        "gross": gross,
        "net": net,
        "special_dates_detected": expenses["special_dates"],
        "shooting_records": shooting_records,
        "data_source": "xlsx" if xlsx_path else ("csv" if use_csv else "api"),
    }

    return result


# ==================== 出力（既存維持） ====================
def print_report(result: dict):
    """給与明細をターミナルに表示"""
    source = result.get("data_source", "csv").upper()
    print(f"\n  [データソース: {source}]")

    print(f"\n【勤務実績】")
    print(f"{'日付':<12} {'出勤':<6} {'退勤':<6} {'実働':<6} {'区分':<12} {'日給':>8} {'車両手当':>8}")
    print("-" * 65)
    for d in result["work_days"]:
        print(f"{d['date']:<12} {d['time_in']:<6} {d['time_out']:<6} "
              f"{d['worked_hours']:>4.1f}h  {d['day_type']:<12} "
              f"{d['wage']:>8,} {d['vehicle']:>8,}")

    # 撮影実績の表示（API取得時のみ）
    if result.get("shooting_records"):
        print(f"\n【CRM撮影実績】")
        for rec in result["shooting_records"]:
            solo = "単独撮影" if rec["is_solo"] else f"同日{rec['same_day_count']}件"
            print(f"  {rec['date']} | {rec['case_name'][:30]} | {solo}")

    print(f"\n【給与明細】 {result['month'][:2]}年{result['month'][2:]}月度")
    print(f"支払予定日: {result['pay_date']}  扶養親族: {result['dependents']}名")
    print("-" * 40)
    print(f"I.  基本給与                {result['base_wage']:>10,} 円")
    print(f"II. 車両手当                {result['vehicle_total']:>10,} 円")
    print(f"III.ガソリン代({result['gas_km']}km x {GAS_RATE}円) {result['gas_amount']:>10,} 円")

    if result["fixed_expenses"]:
        print(f"IV. 実費精算")
        for e in result["fixed_expenses"]:
            print(f"    {e['date']} {e['content'][:15]:<15} {e['amount']:>8,} 円")
        print(f"    小計                  {result['fixed_total']:>10,} 円")

    print("-" * 40)
    print(f"    総支給額                {result['gross']:>10,} 円")
    print(f"V.  源泉徴収税（控除）     {result['withholding']:>10,} 円")
    print("=" * 40)
    print(f"    差引支払額             ¥{result['net']:>10,}")
    print("=" * 40)

    if result["special_dates_detected"]:
        print(f"\n  特約業務の可能性がある日（単価確認が必要）:")
        for d in result["special_dates_detected"]:
            print(f"   -> {d}")
        print(f"   ※ --special オプションで単価を指定してください")


def generate_csv(result: dict, output_path: Path):
    """給与明細CSVを出力"""
    rows = []
    month_label = f"20{result['month'][:2]}年{result['month'][2:]}月度"

    rows.append(["新美 光", f"{month_label} 給与明細", "", ""])
    rows.append(["支払予定日", result["pay_date"], f"扶養親族", f"{result['dependents']}名"])
    rows.append([])
    rows.append(["項目", "内訳", "金額（円）", "備考"])
    rows.append(["I. 基本給与", "", result["base_wage"], ""])

    full_days = [d for d in result["work_days"] if d["day_type"] == "フル日"]
    half_days = [d for d in result["work_days"] if d["day_type"] == "半日"]
    special_days = [d for d in result["work_days"] if d["day_type"] not in ["フル日", "半日", "要確認"]]

    if full_days:
        dates = ", ".join(d["date"][5:] for d in full_days)
        rows.append([f"  フル日給 {FULL_DAY_WAGE:,}円x{len(full_days)}日", dates, FULL_DAY_WAGE * len(full_days), ""])
    if half_days:
        dates = ", ".join(d["date"][5:] for d in half_days)
        rows.append([f"  半日給 {HALF_DAY_WAGE:,}円x{len(half_days)}日", dates, HALF_DAY_WAGE * len(half_days), ""])
    if special_days:
        for d in special_days:
            rows.append([f"  {d['day_type']}", d["date"][5:], d["wage"], "特約"])

    rows.append(["II. 車両手当", "", result["vehicle_total"], ""])
    rows.append([f"  フル日 {VEHICLE_ALLOWANCE_FULL}円x{len(full_days)}日", "", VEHICLE_ALLOWANCE_FULL * len(full_days), ""])
    if half_days:
        rows.append([f"  半日 {VEHICLE_ALLOWANCE_HALF}円x{len(half_days)}日", "", VEHICLE_ALLOWANCE_HALF * len(half_days), ""])

    rows.append(["III. ガソリン代（非課税）", f"{result['gas_km']}km x {GAS_RATE}円/km", result["gas_amount"], ""])

    if result["fixed_expenses"]:
        rows.append(["IV. 実費精算（非課税）", "", result["fixed_total"], ""])
        for e in result["fixed_expenses"]:
            rows.append([f"  {e['content']}", e["date"], e["amount"], e["type"]])

    rows.append([])
    rows.append(["総支給額", "", result["gross"], ""])
    rows.append(["V. 源泉徴収税（控除）", f"甲欄・扶養{result['dependents']}名", -result["withholding"], ""])
    rows.append(["差引支払額（振込額）", "", result["net"], ""])

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    print(f"\n  給与明細CSV出力: {output_path}")


def send_notification(result: dict):
    """洋輔さんにLark DMで通知（確認依頼）"""
    smtp_host = os.getenv("SMTP_HOST", "smtp.larksuite.com")
    smtp_port = int(os.getenv("SMTP_PORT", "465"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    notify_email = os.getenv("LARK_EMAIL_NIIMI", smtp_user)

    month_label = f"20{result['month'][:2]}年{result['month'][2:]}月度"
    subject = f"【給与計算完了】新美 光 {month_label}"

    body = f"""新美 光 {month_label} 給与計算が完了しました。

【概要】
総支給額:   ¥{result['gross']:,}
源泉徴収:   ¥{result['withholding']:,}
差引支払額: ¥{result['net']:,}
支払予定日: {result['pay_date']}
データソース: {result.get('data_source', 'csv').upper()}
"""

    if result["special_dates_detected"]:
        body += f"\n特約業務の確認が必要な日:\n"
        for d in result["special_dates_detected"]:
            body += f"  -> {d}\n"
        body += "\n明細CSVを確認の上、特約単価を指定して再実行してください。\n"
        body += "コマンド例:\n"
        body += f"  python3 payroll_calc.py --month {result['month']} "
        body += f"--special '日付:特約種別:金額'\n"

    try:
        msg = MIMEMultipart()
        msg["From"] = smtp_user
        msg["To"] = notify_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, notify_email, msg.as_string())
        print(f"  通知メール送信完了")
    except Exception as e:
        print(f"  メール通知失敗: {e}")


# ==================== month auto解決 ====================
def resolve_month(month_arg: str) -> str:
    """
    --month auto: 実行日が3日なら前月をYYMM形式で返す。
    それ以外はそのまま返す。
    """
    if month_arg.lower() == "auto":
        now = datetime.now()
        # 前月を計算
        if now.month == 1:
            prev_year = now.year - 1
            prev_month = 12
        else:
            prev_year = now.year
            prev_month = now.month - 1
        yymm = f"{prev_year % 100:02d}{prev_month:02d}"
        print(f"[auto] 前月を自動計算: {yymm}")
        return yymm
    return month_arg


# ==================== Google Sheets書き込み ====================
def _find_sa_key() -> Path | None:
    """Google SAキーファイルを探す"""
    for p in GOOGLE_SA_KEY_PATHS:
        if p.exists():
            return p
    return None


def write_to_google_sheets(result: dict, month: str):
    """Google Sheets「給与明細：新美光」に月別タブを追記"""
    if not HAS_GSPREAD:
        print("  [Sheets] gspread未インストール、スキップ")
        return False

    sa_key = _find_sa_key()
    if not sa_key:
        print("  [Sheets] SAキーファイルが見つかりません、スキップ")
        return False

    year = 2000 + int(month[:2])
    mon = int(month[2:])
    tab_name = f"{year}年{mon}月"

    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = SACredentials.from_service_account_file(str(sa_key), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_ID)

        # 既存タブがあれば削除して再作成
        try:
            existing = sh.worksheet(tab_name)
            sh.del_worksheet(existing)
            print(f"  [Sheets] 既存タブ '{tab_name}' を削除")
        except gspread.exceptions.WorksheetNotFound:
            pass

        ws = sh.add_worksheet(title=tab_name, rows=50, cols=10)

        # ヘッダー
        rows = [
            [f"給与明細 新美 光 {tab_name}", "", "", ""],
            [f"支払予定日: {result['pay_date']}", "", f"扶養親族: {result['dependents']}名", ""],
            [],
            ["日付", "区分", "日給", "車両手当"],
        ]

        # 勤務実績
        for d in result["work_days"]:
            rows.append([d["date"], d["day_type"], d["wage"], d["vehicle"]])

        rows.append([])
        rows.append(["項目", "", "金額", ""])
        rows.append(["基本給与", "", result["base_wage"], ""])
        rows.append(["車両手当", "", result["vehicle_total"], ""])
        rows.append([f"ガソリン代（{result['gas_km']}km）", "", result["gas_amount"], ""])

        if result["fixed_expenses"]:
            for e in result["fixed_expenses"]:
                rows.append([f"実費: {e['content']}", e["date"], e["amount"], e["type"]])

        rows.append([])
        rows.append(["総支給額", "", result["gross"], ""])
        rows.append(["源泉徴収税", "", -result["withholding"], ""])
        rows.append(["差引支払額", "", result["net"], ""])
        rows.append([])
        rows.append(["ステータス", "下書き", "", ""])

        ws.update(range_name="A1", values=rows)
        print(f"  [Sheets] タブ '{tab_name}' に書き込み完了")
        return True

    except Exception as e:
        print(f"  [Sheets] 書き込みエラー: {e}")
        return False


# ==================== PDF生成 ====================
def _fmt_yen(amount: int) -> str:
    return f"¥{amount:,}"


def generate_pdf(result: dict, month: str) -> Path | None:
    """給与明細PDFを生成"""
    if not HAS_REPORTLAB:
        print("  [PDF] reportlab未インストール、スキップ")
        return None

    year = 2000 + int(month[:2])
    mon = int(month[2:])

    PDF_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = PDF_OUTPUT_DIR / f"給与明細_新美光_{month}.pdf"

    # フォント登録
    font_paths_exist = os.path.exists(FONT_REGULAR_PATH) and os.path.exists(FONT_BOLD_PATH)
    if not font_paths_exist:
        # GitHub Actions等ではWindowsフォントがないのでスキップ
        print(f"  [PDF] フォント未検出（{FONT_REGULAR_PATH}）、スキップ")
        return None

    try:
        pdfmetrics.registerFont(TTFont(FONT_REGULAR, FONT_REGULAR_PATH, subfontIndex=0))
        pdfmetrics.registerFont(TTFont(FONT_BOLD, FONT_BOLD_PATH, subfontIndex=0))
    except Exception as e:
        print(f"  [PDF] フォント登録エラー: {e}")
        return None

    PAGE_W, PAGE_H = A4
    MARGIN_L = 20 * mm
    MARGIN_R = 20 * mm
    MARGIN_T = 20 * mm

    CONTENT_W = PAGE_W - MARGIN_L - MARGIN_R

    # 明細の項目を構築
    items = []
    full_days = [d for d in result["work_days"] if d["day_type"] == "フル日"]
    half_days = [d for d in result["work_days"] if d["day_type"] == "半日"]
    special_days = [d for d in result["work_days"] if d["day_type"] not in ["フル日", "半日", "要確認"]]

    if full_days:
        dates = ", ".join(d["date"][5:] for d in full_days)
        items.append({
            "name": f"フル出勤（{dates}）",
            "quantity": len(full_days), "unit": "日",
            "unit_price": FULL_DAY_WAGE,
        })
    if half_days:
        dates = ", ".join(d["date"][5:] for d in half_days)
        items.append({
            "name": f"半日出勤（{dates}）",
            "quantity": len(half_days), "unit": "日",
            "unit_price": HALF_DAY_WAGE,
        })
    for d in special_days:
        items.append({
            "name": f"{d['day_type']}（{d['date'][5:]}）",
            "quantity": 1, "unit": "日",
            "unit_price": d["wage"],
        })

    if result["vehicle_total"] > 0:
        items.append({
            "name": "車両手当",
            "quantity": len(result["work_days"]), "unit": "日",
            "unit_price": result["vehicle_total"] // max(len(result["work_days"]), 1),
        })

    if result["gas_amount"] > 0:
        items.append({
            "name": f"ガソリン代（{result['gas_km']}km）",
            "quantity": result["gas_km"], "unit": "km",
            "unit_price": GAS_RATE,
        })

    for e in result["fixed_expenses"]:
        items.append({
            "name": f"{e['content']}（{e['date'][5:]}）",
            "quantity": 1, "unit": "式",
            "unit_price": e["amount"],
        })

    period = f"{year}年{mon}月1日 〜 {year}年{mon}月{(datetime(year, mon + 1, 1) - timedelta(days=1)).day if mon < 12 else 31}日"

    data = {
        "payee_name": "新美 光",
        "period": period,
        "payment_date": result["pay_date"],
        "items": items,
        "gross_total": result["gross"],
        "withholding_tax": result["withholding"],
        "net_payment": result["net"],
        "tax_notes": [
            "・給与所得に対する源泉徴収（甲欄）",
            f"・課税対象額: {_fmt_yen(result['taxable'])}",
            f"・扶養親族: {result['dependents']}名",
            f"・源泉徴収税額: {_fmt_yen(result['withholding'])}",
        ],
    }

    # generate_payslip.pyのdraw_payslip関数を再利用
    try:
        sys.path.insert(0, str(PDF_OUTPUT_DIR))
        from generate_payslip import draw_payslip, register_fonts
        c = rl_canvas.Canvas(str(output_path), pagesize=A4)
        c.setTitle(f"給与明細_新美光_{year}{mon:02d}")
        c.setAuthor(COMPANY_INFO["name"])
        draw_payslip(c, data)
        c.save()
        print(f"  [PDF] 生成完了: {output_path}")
        return output_path
    except ImportError:
        # generate_payslip.pyが読み込めない場合は簡易PDF
        c = rl_canvas.Canvas(str(output_path), pagesize=A4)
        c.setTitle(f"給与明細_新美光_{year}{mon:02d}")
        y = PAGE_H - MARGIN_T

        c.setFont(FONT_BOLD, 18)
        c.drawCentredString(PAGE_W / 2, y, "支 払 明 細 書")
        y -= 15 * mm

        c.setFont(FONT_REGULAR, 12)
        c.drawString(MARGIN_L, y, f"新美 光 殿")
        y -= 8 * mm
        c.setFont(FONT_REGULAR, 10)
        c.drawString(MARGIN_L, y, f"対象期間: {period}")
        y -= 6 * mm
        c.drawString(MARGIN_L, y, f"支払日: {result['pay_date']}")
        y -= 12 * mm

        for item in items:
            amt = item["quantity"] * item["unit_price"]
            c.drawString(MARGIN_L, y, f"{item['name']}: {item['quantity']}{item['unit']} x {_fmt_yen(item['unit_price'])} = {_fmt_yen(amt)}")
            y -= 6 * mm

        y -= 6 * mm
        c.setFont(FONT_BOLD, 12)
        c.drawString(MARGIN_L, y, f"総支給額: {_fmt_yen(result['gross'])}")
        y -= 7 * mm
        c.setFont(FONT_REGULAR, 10)
        c.drawString(MARGIN_L, y, f"源泉徴収税: -{_fmt_yen(result['withholding'])}")
        y -= 7 * mm
        c.setFont(FONT_BOLD, 14)
        c.drawString(MARGIN_L, y, f"差引支払額: {_fmt_yen(result['net'])}")

        c.save()
        print(f"  [PDF] 簡易PDF生成完了: {output_path}")
        return output_path
    except Exception as e:
        print(f"  [PDF] 生成エラー: {e}")
        return None


# ==================== Lark Base 給与計算テーブルに下書き作成 ====================
def create_payroll_draft_record(result: dict, month: str):
    """給与計算テーブルに下書きレコードを作成"""
    try:
        config = load_config()
        token = lark_get_token(config)
    except Exception as e:
        print(f"  [Lark Base] トークン取得失敗: {e}")
        return None

    year = 2000 + int(month[:2])
    mon = int(month[2:])

    full_days = len([d for d in result["work_days"] if d["day_type"] == "フル日"])
    half_days = len([d for d in result["work_days"] if d["day_type"] == "半日"])
    solo_days = len([d for d in result["work_days"] if d["day_type"] == "単独撮影"])
    multi_days = len([d for d in result["work_days"] if d.get("same_day_count", 0) > 1])
    point_cloud = len([d for d in result["work_days"] if "点群" in d.get("day_type", "")])

    # 基本報酬の内訳（現場 vs 内業の分離は勤怠データからは判定困難なので全額を現場に計上）
    fields = {
        "対象月": month,
        "対象者": "新美光",
        "ステータス": "下書き",
        "フル出勤日数": full_days,
        "半日出勤日数": half_days,
        "単独撮影日数": solo_days,
        "複数現場加算日数": multi_days,
        "点群処理箇所数": point_cloud,
        "基本報酬_現場": result["base_wage"],
        "基本報酬_内業": 0,
        "車両手当": result["vehicle_total"],
        "ガソリン距離km": result["gas_km"],
        "ガソリン代": result["gas_amount"],
        "高速代": sum(e["amount"] for e in result["fixed_expenses"] if "高速" in e.get("content", "")),
        "駐車場代": sum(e["amount"] for e in result["fixed_expenses"] if "駐車" in e.get("content", "")),
        "公共交通機関費": sum(e["amount"] for e in result["fixed_expenses"] if any(k in e.get("content", "") for k in ["電車", "バス", "交通"])),
        "その他実費": sum(e["amount"] for e in result["fixed_expenses"] if not any(k in e.get("content", "") for k in ["高速", "駐車", "電車", "バス", "交通"])),
        "経費精算合計": result["gas_amount"] + result["fixed_total"],
        "課税対象額": result["taxable"],
        "源泉徴収税": result["withholding"],
        "総支給額": result["gross"],
        "差引支払額": result["net"],
    }

    data = {"fields": fields}
    api_result = lark_api_request(
        f"{LARK_API_BASE}/bitable/v1/apps/{CRM_BASE_TOKEN}/tables/{TABLE_PAYROLL}/records",
        method="POST",
        data=data,
        token=token,
    )

    if api_result and api_result.get("code") == 0:
        record_id = api_result.get("data", {}).get("record", {}).get("record_id", "")
        print(f"  [Lark Base] 下書きレコード作成完了: {record_id}")
        return record_id
    else:
        print(f"  [Lark Base] レコード作成失敗: {api_result}")
        return None


# ==================== Lark Bot DM通知 ====================
def send_lark_dm_notification(result: dict, month: str):
    """國本にLark Bot DMで給与明細作成通知を送信"""
    try:
        config = load_config()
        token = lark_get_token(config)
    except Exception as e:
        print(f"  [Lark DM] トークン取得失敗: {e}")
        return False

    year = 2000 + int(month[:2])
    mon = int(month[2:])

    text = (
        f"{mon}月分給与明細を作成しました。確認してください。\n\n"
        f"対象者: 新美 光\n"
        f"総支給額: ¥{result['gross']:,}\n"
        f"源泉徴収: ¥{result['withholding']:,}\n"
        f"差引支払額: ¥{result['net']:,}\n\n"
        f"確認後、送信コマンド:\n"
        f"  python3 payslip_lark_sender.py --month {month} --send"
    )

    data = {
        "receive_id": KUNIMOTO_OPEN_ID,
        "msg_type": "text",
        "content": json.dumps({"text": text}),
    }
    api_result = lark_api_request(
        f"{LARK_API_BASE}/im/v1/messages?receive_id_type=open_id",
        method="POST",
        data=data,
        token=token,
    )

    if api_result and api_result.get("code") == 0:
        print(f"  [Lark DM] 國本に通知送信完了")
        return True
    else:
        print(f"  [Lark DM] 送信失敗: {api_result}")
        return False


# ==================== メイン ====================
def main():
    parser = argparse.ArgumentParser(description="東海エアサービス 給与計算（Lark API統合版）")
    parser.add_argument("--month", required=True, help="対象月 例: 2602 / auto（前月自動計算）")
    parser.add_argument("--dependents", type=int, default=1, help="扶養人数（デフォルト1）")
    parser.add_argument("--special", action="append", default=[],
                        help="特約業務指定 例: '2026-02-10:単独撮影:15000'")
    parser.add_argument("--no-notify", action="store_true", help="メール通知しない")
    parser.add_argument("--output", help="CSV出力先（省略時は自動）")
    parser.add_argument("--csv", action="store_true", help="従来のCSV入力モード（APIを使わない）")
    parser.add_argument("--xlsx", nargs="?", const="auto", default=None,
                        help="Lark勤怠エクスポートxlsx入力モード（パス指定 or 'auto'でDownloadsから自動検出）")
    parser.add_argument("--api", action="store_true", help="API強制モード（フォールバックしない）")
    parser.add_argument("--dry-run", action="store_true", help="API取得テスト（通知・CSV出力しない）")

    args = parser.parse_args()

    # --month auto 解決
    month = resolve_month(args.month)

    # 特約業務の解析
    special_overrides = {}
    for s in args.special:
        parts = s.split(":")
        if len(parts) == 3:
            date, type_name, amount = parts
            special_overrides[date] = {"type": type_name, "amount": int(amount)}

    # 給与計算実行
    result = calculate_payroll(
        month,
        args.dependents,
        special_overrides,
        use_csv=args.csv,
        xlsx_path=args.xlsx,
        dry_run=args.dry_run,
        force_api=args.api,
    )

    # ターミナル表示
    print_report(result)

    if args.dry_run:
        print(f"\n  [dry-run] 計算完了。CSV出力・通知はスキップ。")
        return

    # CSV出力
    output_dir = BASE_DIR / month
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = Path(args.output) if args.output else output_dir / f"給与明細_新美光_{month}.csv"
    generate_csv(result, output_path)

    # PDF生成
    pdf_path = generate_pdf(result, month)

    # Google Sheets書き込み
    write_to_google_sheets(result, month)

    # Lark Base下書きレコード作成
    create_payroll_draft_record(result, month)

    # Lark Bot DM通知（國本に確認依頼）
    send_lark_dm_notification(result, month)

    # メール通知（レガシー、--no-notifyでスキップ推奨）
    if not args.no_notify:
        send_notification(result)


if __name__ == "__main__":
    main()
